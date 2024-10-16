import openpyxl
try:
    wb = openpyxl.load_workbook('produtos.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Nome'
    ws['B1'] = 'Valor'
    ws['C1'] = 'Descrição'
    ws['D1'] = 'Disponível'
def cadastrar_produto():
    nome = input('Digite o nome do produto: ')
    descricao = input('Digite a descrição do produto: ')
    valor = float(input('Digite o valor do produto: '))
    disponivel = input('Digite se o produto está disponível para venda (sim/não): ')
    ws.append([nome, valor, descricao, disponivel])
    wb.save('produtos.xlsx')
    ordenar_produtos()
def ordenar_produtos():
    produtos = list(ws.iter_rows(min_row=2, values_only=True))
    produtos.sort(key=lambda x: float('inf') if x[1] is None else x[1])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    for i, produto in enumerate(produtos, start=2):
        for j, value in enumerate(produto):
            ws.cell(row=i, column=j + 1, value=value)
    wb.save('produtos.xlsx')
def listar_produtos():
    print('Listagem de produtos:')
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f'Nome: {row[0]}, Valor: {row[1]}, Descrição: {row[2]}, Disponível: {row[3]}')
while True:
    print('1. Cadastrar novo produto')
    print('2. Listar produtos')
    print('3. Sair')
    opcao = input('Digite a opção: ')
    if opcao == '1':
        cadastrar_produto()
        listar_produtos()
    elif opcao == '2':
        listar_produtos()
    elif opcao == '3':
        break
    else:
        print('Opção inválida. Tente novamente.')